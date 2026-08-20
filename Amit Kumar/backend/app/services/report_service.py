"""
Reports Service — Textile Waste Intelligence Platform (Milestone 4)
Generates real PDF and Excel reports using reportlab and openpyxl.
All data is fetched from the actual database.
"""
import io
import os
import logging
from datetime import datetime
from typing import Optional

logger = logging.getLogger("twip.reports")


# ══════════════════════════════════════════════════════════════════════════════
#  PDF Report Generation using reportlab
# ══════════════════════════════════════════════════════════════════════════════
def generate_pdf_report(report_type: str, data: dict) -> bytes:
    """Generate a professional PDF report and return bytes."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    # ── Colour palette ──────────────────────────────────────────────────────
    PRIMARY  = colors.HexColor("#10b981")   # emerald
    DARK_BG  = colors.HexColor("#0f172a")   # slate-900
    CARD_BG  = colors.HexColor("#1e293b")   # slate-800
    TEXT_MAIN= colors.HexColor("#f1f5f9")   # slate-100
    TEXT_MUTED=colors.HexColor("#94a3b8")   # slate-400
    ACCENT   = colors.HexColor("#3b82f6")   # blue-500
    WARN     = colors.HexColor("#f59e0b")   # amber-500

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", fontSize=22, textColor=PRIMARY, spaceAfter=6,
                                  fontName="Helvetica-Bold", alignment=TA_CENTER)
    subtitle_style = ParagraphStyle("Subtitle", fontSize=11, textColor=TEXT_MUTED,
                                     spaceAfter=4, fontName="Helvetica", alignment=TA_CENTER)
    section_style = ParagraphStyle("Section", fontSize=13, textColor=PRIMARY,
                                    spaceBefore=12, spaceAfter=6, fontName="Helvetica-Bold")
    body_style = ParagraphStyle("Body", fontSize=9, textColor=TEXT_MAIN,
                                 fontName="Helvetica", spaceAfter=4, leading=14)
    footer_style = ParagraphStyle("Footer", fontSize=7, textColor=TEXT_MUTED,
                                   fontName="Helvetica", alignment=TA_CENTER)

    content = []

    # ── Header ───────────────────────────────────────────────────────────────
    content.append(Paragraph("🌿 Textile Waste Intelligence Platform", title_style))
    content.append(Paragraph(report_type.replace("_", " ").title() + " Report", subtitle_style))
    content.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}", subtitle_style))
    content.append(HRFlowable(width="100%", thickness=2, color=PRIMARY, spaceAfter=10))

    # ── KPI Summary Table ────────────────────────────────────────────────────
    content.append(Paragraph("Executive Summary", section_style))
    kpis = data.get("kpis", [])
    if kpis:
        kpi_data = [["Metric", "Value", "Unit"]]
        for k in kpis:
            kpi_data.append([k.get("label", ""), str(k.get("value", "")), k.get("unit", "")])

        kpi_table = Table(kpi_data, colWidths=[90*mm, 50*mm, 30*mm])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  PRIMARY),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0),  10),
            ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
            ("ALIGN",         (1, 0), (1, -1),  "RIGHT"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [CARD_BG, colors.HexColor("#263248")]),
            ("TEXTCOLOR",     (0, 1), (-1, -1), TEXT_MAIN),
            ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",      (0, 1), (-1, -1), 9),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#334155")),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ]))
        content.append(kpi_table)
        content.append(Spacer(1, 8))

    # ── Inventory Details ────────────────────────────────────────────────────
    inventory = data.get("inventory", [])
    if inventory:
        content.append(Paragraph("Waste Inventory Details", section_style))
        inv_data = [["Batch ID", "Fabric Type", "Qty (kg)", "Condition", "Classification", "Score"]]
        for item in inventory[:30]:  # Limit to 30 rows in PDF
            inv_data.append([
                item.get("waste_batch_id", ""),
                item.get("fabric_type", ""),
                str(item.get("quantity_kg", "")),
                item.get("condition", ""),
                item.get("classification", ""),
                str(item.get("sustainability_score", "")),
            ])
        inv_table = Table(inv_data, colWidths=[32*mm, 28*mm, 20*mm, 22*mm, 28*mm, 18*mm])
        inv_table.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  ACCENT),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0),  8),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [CARD_BG, colors.HexColor("#263248")]),
            ("TEXTCOLOR",     (0, 1), (-1, -1), TEXT_MAIN),
            ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",      (0, 1), (-1, -1), 7.5),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#334155")),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("ALIGN",         (2, 0), (2, -1),  "RIGHT"),
            ("ALIGN",         (5, 0), (5, -1),  "RIGHT"),
        ]))
        content.append(inv_table)
        content.append(Spacer(1, 8))

    # ── Additional sections by report type ──────────────────────────────────
    extra_sections = data.get("sections", [])
    for sec in extra_sections:
        content.append(Paragraph(sec.get("title", ""), section_style))
        content.append(Paragraph(sec.get("body", ""), body_style))
        content.append(Spacer(1, 4))

    # ── Footer ───────────────────────────────────────────────────────────────
    content.append(Spacer(1, 20))
    content.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#334155")))
    content.append(Spacer(1, 4))
    content.append(Paragraph(
        "Generated by TWIP — Textile Waste Intelligence Platform  •  Confidential  •  EcoTextile India",
        footer_style
    ))

    doc.build(content)
    buffer.seek(0)
    return buffer.read()


# ══════════════════════════════════════════════════════════════════════════════
#  Excel Report Generation using openpyxl
# ══════════════════════════════════════════════════════════════════════════════
def generate_excel_report(report_type: str, data: dict) -> bytes:
    """Generate a styled Excel workbook and return bytes."""
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()

    # ── Colour helpers ──────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="10b981")   # emerald header
    ALT_FILL  = PatternFill("solid", fgColor="1e293b")   # dark row
    ALT2_FILL = PatternFill("solid", fgColor="263248")   # alternate row
    KPI_FILL  = PatternFill("solid", fgColor="0f172a")   # kpi bg
    WHITE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    BODY_FONT  = Font(color="F1F5F9", name="Calibri", size=10)
    TITLE_FONT = Font(color="10b981", bold=True, name="Calibri", size=14)
    MUTED_FONT = Font(color="94a3b8", name="Calibri", size=9)
    thin = Side(style="thin", color="334155")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"
    ws_sum.sheet_properties.tabColor = "10b981"
    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 22
    ws_sum.column_dimensions["C"].width = 16

    # Title block
    ws_sum.merge_cells("A1:C1")
    ws_sum["A1"] = "TWIP — Textile Waste Intelligence Platform"
    ws_sum["A1"].font = TITLE_FONT
    ws_sum["A1"].alignment = center
    ws_sum["A1"].fill = KPI_FILL

    ws_sum.merge_cells("A2:C2")
    ws_sum["A2"] = report_type.replace("_", " ").title() + " Report"
    ws_sum["A2"].font = Font(color="94a3b8", name="Calibri", size=11)
    ws_sum["A2"].alignment = center
    ws_sum["A2"].fill = KPI_FILL

    ws_sum.merge_cells("A3:C3")
    ws_sum["A3"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
    ws_sum["A3"].font = MUTED_FONT
    ws_sum["A3"].alignment = center
    ws_sum["A3"].fill = KPI_FILL
    ws_sum.row_dimensions[1].height = 28
    ws_sum.row_dimensions[2].height = 22
    ws_sum.row_dimensions[3].height = 18

    # KPI table headers
    row = 5
    for col, label in enumerate(["Metric", "Value", "Unit"], start=1):
        cell = ws_sum.cell(row=row, column=col, value=label)
        cell.font = WHITE_FONT
        cell.fill = HDR_FILL
        cell.alignment = center
        cell.border = border
    ws_sum.row_dimensions[row].height = 20

    kpis = data.get("kpis", [])
    for i, k in enumerate(kpis):
        row += 1
        fill = ALT_FILL if i % 2 == 0 else ALT2_FILL
        for col, val in enumerate([k.get("label", ""), k.get("value", ""), k.get("unit", "")], start=1):
            cell = ws_sum.cell(row=row, column=col, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = left if col == 1 else center
            cell.border = border
        ws_sum.row_dimensions[row].height = 18

    # ── Sheet 2: Inventory ──────────────────────────────────────────────────
    inventory = data.get("inventory", [])
    if inventory:
        ws_inv = wb.create_sheet("Waste Inventory")
        ws_inv.sheet_properties.tabColor = "3b82f6"
        headers = ["#", "Batch ID", "Fabric Type", "Source", "Qty (kg)", "Color",
                   "Condition", "Classification", "Sustainability Score", "Collected"]
        col_widths = [6, 18, 16, 28, 12, 12, 12, 18, 22, 16]
        for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
            ws_inv.column_dimensions[get_column_letter(i)].width = w
            cell = ws_inv.cell(row=1, column=i, value=h)
            cell.font = WHITE_FONT
            cell.fill = PatternFill("solid", fgColor="3b82f6")
            cell.alignment = center
            cell.border = border
        ws_inv.row_dimensions[1].height = 22

        for idx, item in enumerate(inventory, start=1):
            fill = ALT_FILL if idx % 2 == 1 else ALT2_FILL
            row_data = [
                idx,
                item.get("waste_batch_id", ""),
                item.get("fabric_type", ""),
                item.get("source", ""),
                item.get("quantity_kg", 0),
                item.get("color", ""),
                item.get("condition", ""),
                item.get("classification", ""),
                item.get("sustainability_score", 0),
                str(item.get("collection_date", ""))[:10],
            ]
            for col, val in enumerate(row_data, start=1):
                cell = ws_inv.cell(row=idx + 1, column=col, value=val)
                cell.font = BODY_FONT
                cell.fill = fill
                cell.alignment = center if col != 4 else left
                cell.border = border
            ws_inv.row_dimensions[idx + 1].height = 16

    # ── Sheet 3: Environmental Metrics ──────────────────────────────────────
    ws_env = wb.create_sheet("Environmental Metrics")
    ws_env.sheet_properties.tabColor = "f59e0b"
    ws_env.column_dimensions["A"].width = 35
    ws_env.column_dimensions["B"].width = 18
    ws_env.column_dimensions["C"].width = 14

    env_headers = ["Environmental Indicator", "Value", "Unit"]
    for i, h in enumerate(env_headers, start=1):
        cell = ws_env.cell(row=1, column=i, value=h)
        cell.font = WHITE_FONT
        cell.fill = PatternFill("solid", fgColor="f59e0b")
        cell.alignment = center
        cell.border = border
    ws_env.row_dimensions[1].height = 22

    env_data = data.get("environmental", [])
    for idx, item in enumerate(env_data, start=1):
        fill = ALT_FILL if idx % 2 == 1 else ALT2_FILL
        for col, val in enumerate([item.get("name", ""), item.get("value", ""), item.get("unit", "")], start=1):
            cell = ws_env.cell(row=idx + 1, column=col, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = left if col == 1 else center
            cell.border = border
        ws_env.row_dimensions[idx + 1].height = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
