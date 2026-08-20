"""
Milestone 4: Reports & Export System - shared builders.

Two small, reusable helpers so every report route in reports_routes.py
stays a few lines: one row-list -> Excel workbook, one
title/stats/table -> multi-page PDF (built on top of the same
reportlab primitives already used for the per-item report in
analysis_routes.py).
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas

BRAND_GREEN = (0x1f / 255, 0x6b / 255, 0x4d / 255)  # matches --green-700
BRAND_GREEN_DARK = (0x14 / 255, 0x35 / 255, 0x2b / 255)  # --green-900


def build_excel_report(title: str, columns: list[str], rows: list[list], sheet_name: str = "Report") -> io.BytesIO:
    """rows: list of row-lists, values aligned to `columns`."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Report"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="1F6B4D")

    ws.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().isoformat()}Z")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="666666")

    header_row = 4
    header_fill = PatternFill(start_color="1F6B4D", end_color="1F6B4D", fill_type="solid")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for r_idx, row in enumerate(rows, start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for c_idx, col_name in enumerate(columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(r[c_idx - 1])) for r in rows if c_idx - 1 < len(r)])
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len + 2, 10), 40)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_pdf_report(
    title: str,
    subtitle: str,
    stat_cards: list[tuple[str, str]],
    columns: list[str],
    rows: list[list],
    orientation: str = "landscape",
) -> io.BytesIO:
    """stat_cards: list of (label, value) shown as a summary strip.
    columns/rows: the tabular body of the report."""
    buffer = io.BytesIO()
    pagesize = landscape(A4) if orientation == "landscape" else A4
    pdf = canvas.Canvas(buffer, pagesize=pagesize)
    width, height = pagesize
    margin = 36

    def new_page(header_suffix: str = ""):
        pdf.setFillColorRGB(*BRAND_GREEN_DARK)
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(margin, height - margin, title + header_suffix)
        pdf.setFillColorRGB(0.3, 0.34, 0.32)
        pdf.setFont("Helvetica", 9)
        pdf.drawString(margin, height - margin - 16, subtitle)
        pdf.drawRightString(width - margin, height - margin - 16, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
        pdf.setStrokeColorRGB(0.85, 0.85, 0.85)
        pdf.line(margin, height - margin - 22, width - margin, height - margin - 22)
        return height - margin - 42

    y = new_page()

    # Stat cards strip (only on page 1)
    if stat_cards:
        card_w = (width - 2 * margin - (len(stat_cards) - 1) * 10) / len(stat_cards)
        card_h = 42
        x = margin
        for label, value in stat_cards:
            pdf.setFillColorRGB(0.918, 0.969, 0.937)  # --green-100
            pdf.roundRect(x, y - card_h, card_w, card_h, 5, fill=1, stroke=0)
            pdf.setFillColorRGB(*BRAND_GREEN_DARK)
            pdf.setFont("Helvetica-Bold", 12)
            pdf.drawString(x + 8, y - 20, str(value))
            pdf.setFillColorRGB(0.3, 0.34, 0.32)
            pdf.setFont("Helvetica", 7.5)
            pdf.drawString(x + 8, y - 34, label.upper())
            x += card_w + 10
        y -= card_h + 20

    # Table
    n_cols = max(len(columns), 1)
    col_w = (width - 2 * margin) / n_cols
    row_h = 16

    def draw_table_header():
        nonlocal y
        pdf.setFillColorRGB(*BRAND_GREEN)
        pdf.rect(margin, y - row_h + 4, width - 2 * margin, row_h, fill=1, stroke=0)
        pdf.setFillColorRGB(1, 1, 1)
        pdf.setFont("Helvetica-Bold", 8.5)
        for i, col in enumerate(columns):
            pdf.drawString(margin + i * col_w + 4, y - row_h + 9, str(col)[:40])
        y -= row_h

    draw_table_header()
    pdf.setFont("Helvetica", 8)
    for r_idx, row in enumerate(rows):
        if y < margin + row_h:
            pdf.showPage()
            y = new_page(" (continued)")
            draw_table_header()
            pdf.setFont("Helvetica", 8)
        if r_idx % 2 == 0:
            pdf.setFillColorRGB(0.96, 0.965, 0.95)
            pdf.rect(margin, y - row_h + 4, width - 2 * margin, row_h, fill=1, stroke=0)
        pdf.setFillColorRGB(0.11, 0.14, 0.13)
        for i, value in enumerate(row):
            text = "-" if value is None or value == "" else str(value)
            pdf.drawString(margin + i * col_w + 4, y - row_h + 9, text[:45])
        y -= row_h

    if not rows:
        pdf.setFont("Helvetica-Oblique", 9)
        pdf.drawString(margin + 4, y - 12, "No records available for this report yet.")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer
