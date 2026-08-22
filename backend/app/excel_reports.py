"""
Excel export (Milestone 4 - Reports & Export System). Builds a real
multi-sheet workbook from whatever is actually in the database - an empty
workspace produces sheets with headers only, not fabricated rows.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F6F52", end_color="1F6F52", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1B2420")


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)


def _write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    _autosize(ws)
    return ws


def build_full_export_workbook(batches, analyses, sustainability_rows) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    # Cover sheet
    cover = wb.create_sheet("Summary")
    cover["A1"] = "Reloom — Full Data Export"
    cover["A1"].font = TITLE_FONT
    cover["A2"] = f"Generated {datetime.utcnow().strftime('%d %B %Y, %H:%M UTC')}"
    cover["A4"] = "Sheets in this workbook:"
    for i, name in enumerate(["Inventory", "Classifications", "Sustainability"], start=5):
        cover[f"A{i}"] = f"- {name}"
    _autosize(cover)

    # Inventory sheet
    inv_headers = ["Batch code", "Fabric type", "Source", "Source type", "Quantity (kg)",
                   "Color", "Condition", "Category", "Status", "Collection date", "Registered"]
    inv_rows = [
        [b.batch_code, b.fabric_type.value, b.source, b.source_type, b.quantity_kg,
         b.color or "", b.condition.value, b.category.value, b.status.value,
         b.collection_date.isoformat() if b.collection_date else "",
         b.created_at.strftime("%Y-%m-%d") if b.created_at else ""]
        for b in batches
    ]
    _write_sheet(wb, "Inventory", inv_headers, inv_rows)

    # Classifications sheet
    class_headers = ["Batch code", "Declared fabric", "Suggested fabric", "Confidence",
                      "Contamination", "Damage", "Texture", "Recommended category",
                      "Recyclability score", "Analyzed at"]
    class_rows = []
    for a in analyses:
        batch = a.batch
        class_rows.append([
            batch.batch_code if batch else "",
            batch.fabric_type.value if batch else "",
            a.predicted_fabric_type.value if a.predicted_fabric_type else "",
            f"{(a.fabric_confidence or 0) * 100:.0f}%",
            f"{(a.contamination_score or 0) * 100:.0f}%",
            f"{(a.damage_score or 0) * 100:.0f}%",
            f"{(a.texture_score or 0) * 100:.0f}%",
            a.recommended_category.value if a.recommended_category else "",
            a.recyclability_score if a.recyclability_score is not None else "",
            a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else "",
        ])
    _write_sheet(wb, "Classifications", class_headers, class_rows)

    # Sustainability sheet
    sus_headers = ["Batch code", "Fabric type", "Quantity (kg)", "Category",
                    "Recommended pathway", "CO2e avoided (kg)", "Water avoided (L)", "Landfill diverted (kg)"]
    sus_rows = [
        [r["batch_code"], r["fabric_type"], r["quantity_kg"], r["category"],
         r["recommended_pathway"], r["co2_saved_kg"], r["water_saved_liters"], r["landfill_diverted_kg"]]
        for r in sustainability_rows
    ]
    _write_sheet(wb, "Sustainability", sus_headers, sus_rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
