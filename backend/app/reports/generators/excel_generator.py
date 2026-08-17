"""
Excel Report Generator — Milestone 4
=======================================
Generates professional Excel workbooks using openpyxl.

Workbook structure (varies by report type):
- Summary       — Report identity, key metrics
- AI Results    — Detailed AI classification data
- Recycling     — Recycling recommendations
- Sustainability— Sustainability metrics
- Environmental — Environmental impact data
- Circularity   — Circular economy scores
- Metadata      — Generation info, system details
"""

import os
import json
import logging
from datetime import datetime
from typing import Dict, Any, Optional

logger = logging.getLogger(__name__)

REPORT_TYPE_LABELS = {
    "waste_classification": "Waste Classification Report",
    "recycling":            "Recycling Report",
    "sustainability":       "Sustainability Report",
    "environmental_impact": "Environmental Impact Report",
    "circular_economy":     "Circular Economy Report",
    "esg_summary":          "ESG Summary Report",
}


def _try_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
        return openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl>=3.1.0")


# ─── Style Helpers ─────────────────────────────────────────────────────────────

def _brand_fill(hex_color: str):
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold=False, size=10, color="1A1A2E", italic=False):
    from openpyxl.styles import Font
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")


def _align(h="left", v="center", wrap=False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border():
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="D0D8D4")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_row(ws, row: int, headers: list, fill_hex="1a9966"):
    """Write a styled header row."""
    fill = _brand_fill(fill_hex)
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = fill
        cell.font = _font(bold=True, size=10, color="FFFFFF")
        cell.alignment = _align("center")
        cell.border = _border()


def _data_row(ws, row: int, values: list, alt: bool = False):
    """Write a styled data row."""
    fill = _brand_fill("F2FAF6") if alt else _brand_fill("FFFFFF")
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = _font(size=9)
        cell.alignment = _align("left", wrap=True)
        cell.border = _border()


def _kv_section(ws, start_row: int, title: str, pairs: list) -> int:
    """Write a key-value section to the worksheet. Returns next available row."""
    # Section title
    ws.cell(row=start_row, column=1, value=title).font = _font(bold=True, size=11, color="1a9966")
    ws.cell(row=start_row, column=1).fill = _brand_fill("E8F8F0")
    ws.cell(row=start_row, column=1).alignment = _align()
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    start_row += 1

    for i, (key, val) in enumerate(pairs):
        cell_k = ws.cell(row=start_row + i, column=1, value=str(key))
        cell_v = ws.cell(row=start_row + i, column=2, value=str(val) if val is not None else "—")
        cell_k.font = _font(bold=True, size=9)
        cell_v.font = _font(size=9)
        cell_k.alignment = _align()
        cell_v.alignment = _align(wrap=True)
        cell_k.border = _border()
        cell_v.border = _border()
        if i % 2 == 0:
            cell_k.fill = _brand_fill("F6FCF9")
            cell_v.fill = _brand_fill("F6FCF9")
        else:
            cell_k.fill = _brand_fill("FFFFFF")
            cell_v.fill = _brand_fill("FFFFFF")

    return start_row + len(pairs) + 2


def _auto_width(ws, min_w=12, max_w=55):
    """Auto-resize all columns."""
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 3))


def _title_block(ws, report_title: str, report_type: str, user_name: str, org: str, date_str: str):
    """Write the top title block in the first sheet."""
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "WeaveCycle — Textile Waste Intelligence Platform"
    title_cell.font = _font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = _brand_fill("1a9966")
    title_cell.alignment = _align("center")

    ws.merge_cells("A2:F2")
    ws["A2"].value = REPORT_TYPE_LABELS.get(report_type, "Report")
    ws["A2"].font = _font(bold=True, size=12, color="FFFFFF")
    ws["A2"].fill = _brand_fill("0d7a52")
    ws["A2"].alignment = _align("center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:F3")
    ws["A3"].value = report_title
    ws["A3"].font = _font(bold=False, size=10, color="4a5568", italic=True)
    ws["A3"].fill = _brand_fill("E8F8F0")
    ws["A3"].alignment = _align("center")

    meta_items = [
        ("Generated By:", user_name),
        ("Organization:", org),
        ("Generated Date:", date_str),
    ]
    for i, (k, v) in enumerate(meta_items):
        r = 4 + i
        ws.cell(row=r, column=1, value=k).font = _font(bold=True, size=9)
        ws.cell(row=r, column=2, value=v).font = _font(size=9)
        ws.cell(row=r, column=1).fill = _brand_fill("F6FCF9")
        ws.cell(row=r, column=2).fill = _brand_fill("F6FCF9")


# ─── Per-Sheet Builders ───────────────────────────────────────────────────────

def _build_summary_sheet(wb, data: Dict[str, Any]):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    report_type = data.get("report_type", "waste_classification")
    _title_block(ws, data.get("title", "Report"), report_type,
                 data.get("user_name", "N/A"), data.get("organization_name", "N/A"),
                 data.get("generated_date", "")[:10])

    start = 8
    summary_rows = [
        ("Report ID", data.get("report_id", "—")),
        ("Report Type", REPORT_TYPE_LABELS.get(report_type, report_type)),
        ("Prediction ID", data.get("prediction_id", "—")),
        ("User", data.get("user_name", "—")),
        ("Email", data.get("user_email", "—")),
        ("Organization", data.get("organization_name", "—")),
        ("Role", data.get("role_name", "—")),
        ("Generated Date", data.get("generated_date", "—")[:19] if data.get("generated_date") else "—"),
    ]

    # Add type-specific summary metrics
    if report_type == "waste_classification":
        ai = data.get("ai_results", {}) or {}
        summary_rows += [
            ("Material", ai.get("material", "—")),
            ("Waste Category", ai.get("waste_category", "—")),
            ("Recyclability Score", f"{ai.get('recyclability_score', 0):.1f}%"),
            ("AI Confidence", f"{ai.get('overall_confidence', 0):.1f}%"),
            ("Overall Rating", ai.get("overall_rating", "—")),
        ]
    elif report_type == "recycling":
        rec = data.get("recycling", {}) or {}
        summary_rows += [
            ("Recommended Method", rec.get("recommended_method", "—")),
            ("Material Recovery", f"{rec.get('material_recovery_pct', 0):.1f}%"),
            ("Success Rate", rec.get("success_rate", "—")),
            ("Recovery Difficulty", rec.get("recovery_difficulty", "—")),
        ]
    elif report_type == "sustainability":
        sus = data.get("sustainability", {}) or {}
        summary_rows += [
            ("Sustainability Score", f"{sus.get('sustainability_score', 0):.1f}/100"),
            ("Rating", sus.get("sustainability_rating", "—")),
            ("Resource Recovery Score", f"{sus.get('resource_recovery_score', 0):.1f}/100"),
        ]
    elif report_type == "environmental_impact":
        env = data.get("environmental", {}) or {}
        summary_rows += [
            ("CO₂ Saved", f"{env.get('co2_saved', 0):.2f} kg"),
            ("Water Saved", f"{env.get('water_saved', 0):.0f} L"),
            ("Landfill Diverted", f"{env.get('landfill_diversion', 0):.2f} kg"),
        ]
    elif report_type == "circular_economy":
        circ = data.get("circularity", {}) or {}
        summary_rows += [
            ("Circularity Score", f"{circ.get('circularity_score', 0):.1f}/100"),
            ("Overall Rating", circ.get("overall_rating", "—")),
            ("Classification", circ.get("classification", "—")),
        ]

    _kv_section(ws, start, "Report Summary", summary_rows)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40


def _build_ai_results_sheet(wb, data: Dict[str, Any]):
    ai = data.get("ai_results", {}) or {}
    if not ai:
        return

    ws = wb.create_sheet("AI Results")
    ws.sheet_view.showGridLines = False

    rows = [
        ("Material", ai.get("material")),
        ("Material Confidence", f"{ai.get('material_confidence', 0):.1f}%"),
        ("Fabric Category", ai.get("fabric_category")),
        ("Detected Color", ai.get("detected_color")),
        ("Waste Category", ai.get("waste_category")),
        ("Waste Confidence", f"{ai.get('waste_confidence', 0):.1f}%"),
        ("Recyclability Score", f"{ai.get('recyclability_score', 0):.1f}%"),
        ("Reuse Potential", f"{ai.get('reuse_potential', 0):.1f}%"),
        ("Recovery Difficulty", ai.get("recovery_difficulty")),
        ("Material Recovery Score", f"{ai.get('material_recovery_score', 0):.1f}%"),
        ("Overall Rating", ai.get("overall_rating")),
        ("Overall Confidence", f"{ai.get('overall_confidence', 0):.1f}%"),
        ("Model Version", ai.get("model_version", "v1.0.0")),
        ("Recyclable", "Yes" if ai.get("is_recyclable") else "No"),
        ("Reusable", "Yes" if ai.get("is_reusable") else "No"),
        ("Repairable", "Yes" if ai.get("is_repairable") else "No"),
        ("Hazardous", "Yes" if ai.get("is_hazardous") else "No"),
        ("Mixed Material", "Yes" if ai.get("is_mixed") else "No"),
        ("Contamination Status", ai.get("contamination_status")),
        ("Damage Detection", ai.get("damage_detection")),
        ("Image Quality", ai.get("image_quality")),
    ]
    _kv_section(ws, 1, "AI Classification Results", rows)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35


def _build_recycling_sheet(wb, data: Dict[str, Any]):
    rec = data.get("recycling", {}) or {}
    if not rec:
        return

    ws = wb.create_sheet("Recycling")
    ws.sheet_view.showGridLines = False

    main_rows = [
        ("Recommended Method", rec.get("recommended_method")),
        ("Technique", rec.get("technique")),
        ("Recovery Recommendation", rec.get("recovery_recommendation")),
        ("Recovery Difficulty", rec.get("recovery_difficulty")),
        ("Material Recovery %", f"{rec.get('material_recovery_pct', 0):.1f}%"),
        ("Success Rate", rec.get("success_rate")),
        ("Estimated Cost", rec.get("estimated_cost")),
        ("Estimated Time", rec.get("estimated_time")),
        ("Environmental Benefit", rec.get("environmental_benefit")),
        ("Industry Applications", rec.get("industry_applications")),
    ]
    next_row = _kv_section(ws, 1, "Primary Recycling Recommendation", main_rows)

    # All recommendations table
    all_recs = rec.get("all_recommendations", []) or []
    if all_recs:
        headers = ["Method", "Priority", "Difficulty", "Success Rate", "Cost", "Time", "Expected Output"]
        _header_row(ws, next_row, headers)
        for i, r in enumerate(all_recs, 1):
            _data_row(ws, next_row + i, [
                r.get("method", "—"),
                r.get("priority", "—"),
                r.get("difficulty", "—"),
                r.get("success_rate", "—"),
                r.get("estimated_cost", "—"),
                r.get("estimated_time", "—"),
                r.get("expected_output", "—"),
            ], alt=i % 2 == 0)

    _auto_width(ws)


def _build_sustainability_sheet(wb, data: Dict[str, Any]):
    sus = data.get("sustainability", {}) or {}
    if not sus:
        return

    ws = wb.create_sheet("Sustainability")
    ws.sheet_view.showGridLines = False

    rows = [
        ("Sustainability Score", f"{sus.get('sustainability_score', 0):.1f} / 100"),
        ("Environmental Benefit Score", f"{sus.get('environmental_benefit_score', 0):.1f} / 100"),
        ("Resource Recovery Score", f"{sus.get('resource_recovery_score', 0):.1f} / 100"),
        ("Material Longevity Score", f"{sus.get('material_longevity_score', 0):.1f} / 100"),
        ("Waste Diversion Score", f"{sus.get('waste_diversion_score', 0):.1f} / 100"),
        ("Carbon Footprint", sus.get("carbon_footprint")),
        ("Sustainability Rating", sus.get("sustainability_rating")),
    ]
    next_row = _kv_section(ws, 1, "Sustainability Metrics", rows)

    insights = sus.get("insights", []) or []
    if insights:
        ws.cell(row=next_row, column=1, value="AI Insights").font = _font(bold=True, size=11, color="1a9966")
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=2)
        next_row += 1
        for i, ins in enumerate(insights):
            cell = ws.cell(row=next_row + i, column=1, value=f"• {ins}")
            cell.font = _font(size=9)
            ws.merge_cells(start_row=next_row + i, start_column=1, end_row=next_row + i, end_column=2)

    _auto_width(ws)


def _build_environmental_sheet(wb, data: Dict[str, Any]):
    env = data.get("environmental", {}) or {}
    if not env:
        return

    ws = wb.create_sheet("Environmental")
    ws.sheet_view.showGridLines = False

    rows = [
        ("CO₂ Saved", f"{env.get('co2_saved', 0):.2f} kg"),
        ("Water Saved", f"{env.get('water_saved', 0):.0f} Liters"),
        ("Energy Saved", f"{env.get('energy_saved', 0):.1f} kWh"),
        ("Landfill Diversion", f"{env.get('landfill_diversion', 0):.2f} kg"),
        ("Resource Conservation", f"{env.get('resource_conservation', 0):.2f} kg"),
        ("Equivalent Trees Planted", f"{env.get('equivalent_trees', 0):.1f}"),
        ("Equivalent kWh Electricity", f"{env.get('equivalent_electricity', 0):.1f} kWh"),
        ("Equivalent Water Bottles", f"{env.get('equivalent_water_bottles', 0):.0f}"),
        ("Equivalent Household Energy Days", f"{env.get('equivalent_household_energy', 0):.1f}"),
    ]
    _kv_section(ws, 1, "Environmental Impact Estimation", rows)
    _auto_width(ws)


def _build_circularity_sheet(wb, data: Dict[str, Any]):
    circ = data.get("circularity", {}) or {}
    if not circ:
        return

    ws = wb.create_sheet("Circularity")
    ws.sheet_view.showGridLines = False

    rows = [
        ("Circularity Score", f"{circ.get('circularity_score', 0):.1f} / 100"),
        ("Reuse Potential", f"{circ.get('reuse_potential', 0):.1f}%"),
        ("Recovery Efficiency", f"{circ.get('recovery_efficiency', 0):.1f}%"),
        ("Material Retention", f"{circ.get('material_retention', 0):.1f}%"),
        ("Lifecycle Extension", f"{circ.get('lifecycle_extension', 0):.1f}%"),
        ("Circularity Index", f"{circ.get('circularity_index', 0):.3f}"),
        ("Classification", circ.get("classification")),
        ("Overall Rating", circ.get("overall_rating")),
    ]
    _kv_section(ws, 1, "Circular Economy Metrics", rows)
    _auto_width(ws)


def _build_metadata_sheet(wb, data: Dict[str, Any]):
    ws = wb.create_sheet("Metadata")
    ws.sheet_view.showGridLines = False

    rows = [
        ("Platform", "WeaveCycle — Textile Waste Intelligence Platform"),
        ("Milestone", "4 — Reports & Export System"),
        ("Report ID", data.get("report_id", "—")),
        ("Report Type", REPORT_TYPE_LABELS.get(data.get("report_type", ""), data.get("report_type", "—"))),
        ("Report Title", data.get("title", "—")),
        ("Generated By", data.get("user_name", "—")),
        ("User Email", data.get("user_email", "—")),
        ("Organization", data.get("organization_name", "—")),
        ("Role", data.get("role_name", "—")),
        ("Generated At (UTC)", data.get("generated_date", "—")[:19] if data.get("generated_date") else "—"),
        ("Prediction ID", data.get("prediction_id", "—")),
        ("AI Model Version", "EfficientNet-B0 v1.0.0"),
        ("Export Format", "Microsoft Excel (.xlsx)"),
    ]
    _kv_section(ws, 1, "Report Metadata", rows)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50


# ─── Main Entry Point ─────────────────────────────────────────────────────────

def generate_excel(report_data: Dict[str, Any], output_dir: str) -> str:
    """
    Generate an Excel workbook and save to output_dir.
    Returns the saved file path.
    """
    openpyxl = _try_openpyxl()

    report_type = report_data.get("report_type", "waste_classification")
    report_id = report_data.get("report_id", "UNKNOWN")
    report_title = report_data.get("title", "WeaveCycle Report")

    safe_title = "".join(c if c.isalnum() or c in "-_ " else "_" for c in report_title)[:50]
    filename = f"{report_type}_{safe_title}_{report_id[:8]}.xlsx".replace(" ", "_")
    output_path = os.path.join(output_dir, filename)

    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()

    # Always build Summary and Metadata
    _build_summary_sheet(wb, report_data)
    _build_metadata_sheet(wb, report_data)

    # Build type-specific sheets
    _build_ai_results_sheet(wb, report_data)

    if report_type in ("recycling",):
        _build_recycling_sheet(wb, report_data)
    if report_type == "sustainability":
        _build_sustainability_sheet(wb, report_data)
    if report_type == "environmental_impact":
        _build_environmental_sheet(wb, report_data)
    if report_type == "circular_economy":
        _build_circularity_sheet(wb, report_data)
    if report_type == "esg_summary":
        _build_esg_sheet(wb, report_data)

    # Include all sections when available regardless of type
    if report_data.get("recycling") and report_type != "recycling":
        _build_recycling_sheet(wb, report_data)
    if report_data.get("sustainability") and report_type != "sustainability":
        _build_sustainability_sheet(wb, report_data)
    if report_data.get("environmental") and report_type != "environmental_impact":
        _build_environmental_sheet(wb, report_data)
    if report_data.get("circularity") and report_type != "circular_economy":
        _build_circularity_sheet(wb, report_data)

    wb.save(output_path)
    logger.info(f"Excel generated: {output_path}")
    return output_path


def _build_esg_sheet(wb, data: Dict[str, Any]):
    esg = data.get("esg", {}) or {}
    if not esg:
        return

    ws = wb.create_sheet("ESG Summary")
    ws.sheet_view.showGridLines = False

    # Title & Exec Summary
    ws.cell(row=1, column=1, value="ESG Executive Summary").font = _font(bold=True, size=11, color="1a9966")
    ws.cell(row=1, column=1).fill = _brand_fill("E8F8F0")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    
    cell_exec = ws.cell(row=2, column=1, value=esg.get("executive_summary", "—"))
    cell_exec.font = _font(size=9)
    cell_exec.alignment = _align(wrap=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    
    # Pillar Details
    rows_e = [
        ("ESG Score", f"{esg.get('esg_score', 0):.1f}"),
        ("ESG Rating", esg.get("esg_rating", "—")),
        ("Sustainability Score", f"{esg.get('sustainability_score', 0):.1f} / 100"),
        ("Environmental Rating", esg.get("sustainability_rating", "—")),
        ("Carbon Footprint", esg.get("carbon_footprint", "—")),
        ("CO₂ Savings", f"{esg.get('co2_saved', 0):.2f} kg"),
        ("Water Savings", f"{esg.get('water_saved', 0):.0f} L"),
        ("Landfill Diversion", f"{esg.get('landfill_diversion', 0):.2f} kg"),
        ("Resource Recovery Score", f"{esg.get('resource_recovery_score', 0):.1f}"),
        ("Circularity Score", f"{esg.get('circularity_score', 0):.1f}"),
        ("Waste Diversion Score", f"{esg.get('waste_diversion_score', 0):.1f}"),
        ("Recycling Recommendation", esg.get("recycling_recommendation", "—")),
        ("Material Recovery Score", f"{esg.get('material_recovery_score', 0):.1f}%"),
    ]
    next_row = _kv_section(ws, 4, "Environmental (E) Metrics", rows_e)
    
    rows_s = [
        ("Compliance Status", esg.get("compliance_status", "Not Available")),
        ("Waste Handling Safety", esg.get("waste_handling_safety", "Not Available")),
        ("Hazardous Material Detection", esg.get("hazardous_material_detection", "Not Available")),
        ("Contamination Risk", esg.get("contamination_risk", "Not Available")),
        ("Supply Chain Transparency", esg.get("supply_chain_transparency", "Not Available")),
    ]
    next_row = _kv_section(ws, next_row + 1, "Social (S) Indicators", rows_s)

    rows_g = [
        ("Prediction Confidence", f"{esg.get('prediction_confidence', 0):.1f}%" if esg.get("prediction_confidence") else "—"),
        ("AI Model Version", esg.get("model_version", "Not Available")),
        ("Dataset Used", esg.get("dataset_used", "Not Available")),
        ("Report Generated By", esg.get("generated_by", "—")),
        ("Report Generated On", esg.get("generated_on", "—")),
        ("Prediction ID", esg.get("prediction_id", "—")),
        ("Waste Batch ID", esg.get("waste_batch_id", "—")),
        ("Dataset Traceability", esg.get("dataset_traceability", "Not Available")),
        ("Audit Timestamp", esg.get("audit_timestamp", "—")),
    ]
    _kv_section(ws, next_row + 1, "Governance (G) Metadata", rows_g)

    _auto_width(ws)
