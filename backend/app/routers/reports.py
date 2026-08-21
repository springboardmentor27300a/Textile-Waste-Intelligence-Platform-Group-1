import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import current_user
from ..models import User, WasteBatch
from .dashboard import esg as esg_block
from .inventory import _visible

router = APIRouter(prefix="/api/reports", tags=["reports"])

INK = colors.HexColor("#12211C")
INDIGO = colors.HexColor("#24406B")
RULE = colors.HexColor("#C6CCC7")

# Slugs used in the URL for each individual report, mapped to the title
# shown on the document and the short note printed under it.
REPORT_META = {
    "classification": {
        "title": "Waste classification report",
        "note": "Material and waste class per batch, with confidence.",
    },
    "recycling": {
        "title": "Recycling report",
        "note": "Recommended route and fit score per batch.",
    },
    "sustainability": {
        "title": "Sustainability report",
        "note": "Diversion rate, circularity and ESG position.",
    },
    "environmental": {
        "title": "Environmental impact report",
        "note": "CO\u2082, water, landfill and virgin-fibre savings.",
    },
    "circular-economy": {
        "title": "Circular economy report",
        "note": "Mass by recovery route across the facility.",
    },
}


def _rows(db: Session, user: User):
    out = []
    for batch in _visible(db, user).order_by(WasteBatch.created_at.desc()).all():
        analysis = batch.analyses[0] if batch.analyses else None
        top_reco = analysis.recommendations[0] if analysis and analysis.recommendations else None
        out.append({
            "Batch": batch.batch_code,
            "Source": batch.source,
            "Quantity (kg)": round(batch.quantity_kg, 1),
            "Condition": batch.condition,
            "Status": batch.status.value,
            "Material": analysis.material if analysis else "—",
            "Confidence": round(analysis.material_confidence, 3) if analysis else "—",
            "Waste category": analysis.waste_category.value if analysis else "—",
            "Recyclability": analysis.recyclability_score if analysis else "—",
            "Reuse": analysis.reuse_score if analysis else "—",
            "Circularity": analysis.circularity_score if analysis else "—",
            "Band": analysis.circularity_band if analysis else "—",
            "Route": top_reco["route"] if top_reco else "—",
            "Fit score": top_reco["fit"] if top_reco else "—",
            "CO2 saved (kg)": analysis.environmental_impact.get("co2_saved_kg", 0) if analysis else "—",
            "Water saved (L)": analysis.environmental_impact.get("water_saved_litres", 0) if analysis else "—",
        })
    return out


def _styles():
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=20, textColor=INK, alignment=0)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, textColor=INK)
    eyebrow = ParagraphStyle("eyebrow", parent=styles["Normal"], fontSize=8,
                             textColor=INDIGO, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=9, textColor=INK)
    return h1, h2, eyebrow, body


def _document(filename_prefix: str):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=18 * mm, bottomMargin=16 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            title=filename_prefix)
    return buffer, doc


def _cover(title: str, user: User, esg: dict):
    h1, h2, eyebrow, body = _styles()
    return [
        Paragraph(f"{esg['reporting_period'].upper()} &nbsp;·&nbsp; "
                  f"{user.organisation or user.full_name}", eyebrow),
        Paragraph(title, h1),
        Spacer(1, 6 * mm),
    ]


def _register_table(rows, columns, col_widths, limit=40):
    _, h2, _, _ = _styles()
    data = [columns] + [[str(r[c]) for c in columns] for r in rows[:limit]]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 0), (-1, 0), INDIGO),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F3F1")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, RULE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def _headline_table(esg: dict):
    headline = [
        ["Waste diversion rate", f"{esg['waste_diversion_rate']}%"],
        ["Landfill avoided", f"{esg['landfill_avoided_kg']:,.0f} kg"],
        ["CO\u2082 saved", f"{esg['co2_saved_tonnes']:,.2f} t"],
        ["Water saved", f"{esg['water_saved_kilolitres']:,.1f} kL"],
        ["Virgin fibre replaced", f"{esg['virgin_fibre_replaced_kg']:,.0f} kg"],
        ["Hazardous batches", str(esg["hazardous_batches"])],
    ]
    table = Table(headline, colWidths=[70 * mm, 40 * mm])
    table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#5A6560")),
        ("TEXTCOLOR", (1, 0), (1, -1), INK),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, RULE),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    return table


def _routes_table(esg: dict):
    data = [["Recovery route", "Mass (kg)"]] + [
        [r["route"], f"{r['kg']:,.1f}"] for r in esg["recovery_routes"]
    ]
    table = Table(data, colWidths=[80 * mm, 40 * mm])
    table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 0), (-1, 0), INDIGO),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F3F1")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, RULE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return table


def _build_section(kind: str, rows, esg: dict, body_style):
    """Return the story elements (minus the cover) for one report kind."""
    story = []

    if kind == "classification":
        story += [Paragraph("Batch register — classification", body_style), Spacer(1, 2 * mm),
                  _register_table(rows, ["Batch", "Material", "Confidence", "Waste category",
                                          "Quantity (kg)"],
                                   [30 * mm, 30 * mm, 24 * mm, 42 * mm, 28 * mm]),
                  Spacer(1, 6 * mm),
                  Paragraph("Confidence is the model's certainty in the assigned material label. "
                            "Waste category follows the facility's classification taxonomy.",
                            body_style)]

    elif kind == "recycling":
        story += [Paragraph("Batch register — recycling routes", body_style), Spacer(1, 2 * mm),
                  _register_table(rows, ["Batch", "Material", "Quantity (kg)", "Route", "Fit score"],
                                   [28 * mm, 28 * mm, 24 * mm, 45 * mm, 24 * mm]),
                  Spacer(1, 6 * mm),
                  Paragraph("Route is the top-ranked recovery option per batch. Fit score reflects "
                            "how well the batch's material and condition suit that route.",
                            body_style)]

    elif kind == "sustainability":
        story += [_headline_table(esg), Spacer(1, 8 * mm),
                  Paragraph("Batch register — circularity", body_style), Spacer(1, 2 * mm),
                  _register_table(rows, ["Batch", "Material", "Quantity (kg)", "Circularity", "Band"],
                                   [28 * mm, 28 * mm, 24 * mm, 24 * mm, 45 * mm]),
                  Spacer(1, 6 * mm),
                  Paragraph("Diversion rate is the share of intake mass routed away from landfill. "
                            "Figures are calculated across every batch visible to your account.",
                            body_style)]

    elif kind == "environmental":
        story += [_headline_table(esg), Spacer(1, 8 * mm),
                  Paragraph("Batch register — environmental savings", body_style), Spacer(1, 2 * mm),
                  _register_table(rows, ["Batch", "Material", "Quantity (kg)",
                                          "CO2 saved (kg)", "Water saved (L)"],
                                   [28 * mm, 28 * mm, 24 * mm, 32 * mm, 32 * mm]),
                  Spacer(1, 6 * mm),
                  Paragraph("Impact figures are modelled from published life-cycle ranges for "
                            "virgin fibre production. Replace them with facility-measured values "
                            "before external reporting.", body_style)]

    elif kind == "circular-economy":
        story += [Paragraph("Mass by recovery route", body_style), Spacer(1, 2 * mm),
                  _routes_table(esg), Spacer(1, 8 * mm),
                  Paragraph("Batch register — routing detail", body_style), Spacer(1, 2 * mm),
                  _register_table(rows, ["Batch", "Material", "Quantity (kg)", "Route", "Band"],
                                   [28 * mm, 28 * mm, 24 * mm, 42 * mm, 31 * mm]),
                  Spacer(1, 6 * mm),
                  Paragraph("Recovery route mass is aggregated from each batch's top-ranked "
                            "recommendation.", body_style)]

    return story


@router.get("/excel")
def excel_report(db: Session = Depends(get_db), user: User = Depends(current_user)):
    rows = _rows(db, user)
    if not rows:
        raise HTTPException(status_code=404, detail="Register a batch before exporting a report.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Batches"
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row[h] for h in headers])
    for column, header in enumerate(headers, start=1):
        width = max(len(str(header)), *(len(str(r[header])) for r in rows)) + 3
        sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = min(width, 32)

    summary = workbook.create_sheet("Sustainability")
    for key, value in esg_block(db, user).items():
        if isinstance(value, list):
            summary.append([key.replace("_", " ").title()])
            for item in value:
                summary.append(["", item.get("route", ""), item.get("kg", "")])
        else:
            summary.append([key.replace("_", " ").title(), value])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"textile-waste-report-{datetime.now(timezone.utc):%Y%m%d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/pdf")
def pdf_report(db: Session = Depends(get_db), user: User = Depends(current_user)):
    """Combined report: every section, one after another, in a single PDF."""
    rows = _rows(db, user)
    if not rows:
        raise HTTPException(status_code=404, detail="Register a batch before exporting a report.")
    esg = esg_block(db, user)

    buffer, doc = _document("Textile waste intelligence report")
    _, h2, _, body = _styles()

    story = _cover("Textile waste intelligence report", user, esg)
    story += [_headline_table(esg), Spacer(1, 8 * mm),
              Paragraph("Batch register", h2), Spacer(1, 2 * mm)]

    columns = ["Batch", "Material", "Quantity (kg)", "Waste category", "Circularity", "Route"]
    story.append(_register_table(rows, columns,
                                  [32 * mm, 25 * mm, 22 * mm, 30 * mm, 20 * mm, 35 * mm]))
    story += [Spacer(1, 6 * mm),
              Paragraph("Impact figures are modelled from published life-cycle ranges for virgin "
                        "fibre production and the recovery rate of the recommended route. "
                        "Replace them with facility-measured values before external reporting.",
                        body)]
    doc.build(story)
    buffer.seek(0)
    filename = f"textile-waste-report-{datetime.now(timezone.utc):%Y%m%d}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/pdf/{report_type}")
def single_pdf_report(report_type: str, db: Session = Depends(get_db),
                      user: User = Depends(current_user)):
    """One of the five individual reports, downloaded on its own."""
    meta = REPORT_META.get(report_type)
    if not meta:
        raise HTTPException(status_code=404, detail="Unknown report type.")

    rows = _rows(db, user)
    if not rows:
        raise HTTPException(status_code=404, detail="Register a batch before exporting a report.")
    esg = esg_block(db, user)

    buffer, doc = _document(meta["title"])
    _, _, _, body = _styles()

    story = _cover(meta["title"], user, esg)
    story += _build_section(report_type, rows, esg, body)
    doc.build(story)
    buffer.seek(0)

    slug = report_type.replace("-", "_")
    filename = f"{slug}-report-{datetime.now(timezone.utc):%Y%m%d}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})