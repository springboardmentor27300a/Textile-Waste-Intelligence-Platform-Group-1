from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


class ExcelService:
    """
    =========================================================

        Textile Waste Intelligence Platform

            Excel Export Service

    =========================================================
    """

    HEADER_FILL = PatternFill(
        fill_type="solid",
        start_color="1E3A5F",
        end_color="1E3A5F",
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
    )

    @staticmethod
    def export(
        sheet_name: str,
        rows: list[dict],
    ):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = sheet_name

        if not rows:

            worksheet["A1"] = "No Data Available"

        else:

            headers = list(rows[0].keys())

            for column, header in enumerate(headers, start=1):

                cell = worksheet.cell(
                    row=1,
                    column=column,
                )

                cell.value = header.replace(
                    "_",
                    " ",
                ).title()

                cell.font = ExcelService.HEADER_FONT

                cell.fill = ExcelService.HEADER_FILL

                cell.alignment = Alignment(
                    horizontal="center",
                )

            for row_index, row in enumerate(rows, start=2):

                for column, value in enumerate(
                    row.values(),
                    start=1,
                ):

                    worksheet.cell(
                        row=row_index,
                        column=column,
                    ).value = value

        buffer = BytesIO()

        workbook.save(buffer)

        buffer.seek(0)

        return buffer