"""Dedicated operational, environmental, and circular-economy reports."""

import json
from collections import Counter
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.user import InventoryItem, User
from app.utils.permissions import get_current_user, scope_inventory_query

router = APIRouter(prefix="/api/reports", tags=["reports"])

REPORTS = {
    "waste-classification": {
        "title": "Waste Classification Report",
        "columns": ["Batch", "Fabric", "Classification", "Condition", "Quality", "Quantity (kg)", "Status"],
    },
    "recycling": {
        "title": "Recycling and Recovery Report",
        "columns": ["Batch", "Fabric", "Status", "Recyclability", "Recovery %", "Recoverable kg", "Processing method", "Recommended action"],
    },
    "environmental-impact": {
        "title": "Environmental Impact Report",
        "columns": ["Batch", "Fabric", "Waste kg", "CO2 saved kg", "Water saved L", "Landfill reduced kg", "Environmental score"],
    },
    "circular-economy": {
        "title": "Circular Economy Report",
        "columns": ["Batch", "Fabric", "Circularity", "Category", "Reuse", "Recovery %", "Recoverable kg", "Decision"],
    },
    "esg": {
        "title": "ESG Sustainability Report",
        "columns": ["Batch", "Waste kg", "CO2 saved kg", "Water saved L", "Diversion %", "Sustainability", "Circularity", "Category"],
    },
}


def _analysis(batch):
    try:
        return json.loads(batch.analysis_results or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}


def _number(value):
    return round(float(value or 0), 2)


def _rows(db: Session, user: User, report_type: str):
    batches = scope_inventory_query(db.query(InventoryItem), user).order_by(InventoryItem.id.desc()).all()
    rows = []
    for batch in batches:
        assessment = batch.assessment
        analysis = _analysis(batch)
        classification = analysis.get("waste_classification", {})
        quantity = _number(batch.quantity_kg or (assessment.quantity_kg if assessment else 0))
        if report_type == "waste-classification":
            rows.append([batch.waste_batch_id, batch.fabric_type, classification.get("category", batch.condition), batch.condition, classification.get("quality", "Not analysed"), quantity, batch.status])
        elif report_type == "recycling":
            rows.append([batch.waste_batch_id, batch.fabric_type, batch.status, _number(assessment.recyclability_score if assessment else 0), _number(assessment.material_recovery_score if assessment else 0), _number(assessment.recoverable_material_kg if assessment else 0), assessment.recommended_processing_method if assessment else "Assessment required", assessment.recommended_action if assessment else "Assessment required"])
        elif report_type == "environmental-impact":
            rows.append([batch.waste_batch_id, batch.fabric_type, quantity, _number(assessment.co2_saved_kg if assessment else 0), _number(assessment.water_saved_litres if assessment else 0), _number(assessment.landfill_reduction_kg if assessment else 0), _number(assessment.environmental_benefit_score if assessment else 0)])
        elif report_type == "circular-economy":
            rows.append([batch.waste_batch_id, batch.fabric_type, _number(assessment.circularity_score if assessment else 0), assessment.circularity_category if assessment else "Not assessed", _number(assessment.reuse_score if assessment else 0), _number(assessment.material_recovery_score if assessment else 0), _number(assessment.recoverable_material_kg if assessment else 0), assessment.recommended_action if assessment else "Assessment required"])
        elif report_type == "esg":
            diversion = _number((assessment.landfill_reduction_kg / assessment.quantity_kg * 100) if assessment and assessment.quantity_kg else 0)
            rows.append([batch.waste_batch_id, quantity, _number(assessment.co2_saved_kg if assessment else 0), _number(assessment.water_saved_litres if assessment else 0), diversion, _number(assessment.sustainability_score if assessment else 0), _number(assessment.circularity_score if assessment else 0), assessment.circularity_category if assessment else "Not assessed"])
    return rows


def _validate(report_type: str):
    if report_type not in REPORTS:
        raise HTTPException(status_code=404, detail="Unknown report type")
    return REPORTS[report_type]


@router.get("/{report_type}/pdf")
def download_report_pdf(report_type: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    config = _validate(report_type)
    rows = _rows(db, user, report_type)
    stream = BytesIO()
    document = SimpleDocTemplate(stream, pagesize=landscape(A4), rightMargin=10 * mm, leftMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph(config["title"], styles["Title"]), Paragraph(f"Total batches: {len(rows)}", styles["Normal"]), Spacer(1, 5 * mm)]
    table = Table([config["columns"], *rows], repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")), ("FONTSIZE", (0, 0), (-1, -1), 7), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")])]))
    story.append(table)
    document.build(story)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={report_type}-report.pdf"})


@router.get("/{report_type}/excel")
def download_report_excel(report_type: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    config = _validate(report_type)
    rows = _rows(db, user, report_type)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.append(config["columns"])
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="0F766E")
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={report_type}-report.xlsx"})

