"""Downloadable sustainability reports for the current user's accessible batches."""

import csv
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.user import InventoryItem, User
from app.services.assessment_service import assessment_to_dict
from app.services.sustainability_service import aggregate_assessments
from app.utils.permissions import get_current_user, scope_inventory_query

router = APIRouter(prefix="/api/reports/sustainability", tags=["sustainability reports"])


def _report_data(db: Session, user: User):
    batches = scope_inventory_query(db.query(InventoryItem), user).filter(InventoryItem.assessment.has()).order_by(InventoryItem.id.desc()).all()
    assessments = [batch.assessment for batch in batches]
    return aggregate_assessments(assessments), [assessment_to_dict(item) for item in assessments]


@router.get("/csv")
def download_csv(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    _, rows = _report_data(db, user)
    stream = StringIO()
    columns = ["batch_id", "quantity_kg", "co2_saved_kg", "water_saved_litres", "landfill_reduction_kg", "recoverable_material_kg", "recyclability_score", "reuse_score", "material_recovery_score", "sustainability_score", "circularity_score", "circularity_category", "recommended_action", "recommended_processing_method", "updated_at"]
    writer = csv.DictWriter(stream, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return StreamingResponse(iter([stream.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=sustainability-report.csv"})


@router.get("/pdf")
def download_pdf(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    summary, rows = _report_data(db, user)
    stream = BytesIO()
    document = SimpleDocTemplate(stream, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph("Textile Sustainability Report", styles["Title"]), Spacer(1, 5 * mm)]
    metrics = [
        ["Total waste", f'{summary["total_waste_kg"]:,.2f} kg', "CO2 saved", f'{summary["co2_saved_kg"]:,.2f} kg'],
        ["Water saved", f'{summary["water_saved_litres"]:,.2f} L', "Waste diversion", f'{summary["waste_diversion_percentage"]:,.2f}%'],
        ["Recoverable material", f'{summary["recoverable_material_kg"]:,.2f} kg', "Average circularity", f'{summary["average_circularity_score"]:,.2f}/100'],
    ]
    metrics_table = Table(metrics, colWidths=[42 * mm, 35 * mm, 48 * mm, 35 * mm])
    metrics_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ECFDF5")), ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#A7F3D0")), ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"), ("PADDING", (0, 0), (-1, -1), 7)]))
    story.extend([metrics_table, Spacer(1, 6 * mm)])
    data = [["Batch", "Waste kg", "CO2 kg", "Water L", "Recovery kg", "Circularity", "Category", "Recommendation"]]
    for row in rows:
        data.append([row["batch_id"], row["quantity_kg"], row["co2_saved_kg"], row["water_saved_litres"], row["recoverable_material_kg"], row["circularity_score"], row["circularity_category"], row["recommended_action"]])
    table = Table(data, repeatRows=1, colWidths=[25 * mm, 20 * mm, 20 * mm, 24 * mm, 23 * mm, 22 * mm, 48 * mm, 38 * mm])
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")), ("FONTSIZE", (0, 0), (-1, -1), 7), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")])]))
    story.append(table)
    document.build(story)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=sustainability-report.pdf"})


@router.get("/excel")
def download_excel(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    summary, rows = _report_data(db, user)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Summary"
    overview.append(["Sustainability metric", "Value"])
    for key, value in summary.items():
        if key != "category_distribution":
            overview.append([key.replace("_", " ").title(), value])
    overview.append([])
    overview.append(["Circularity category", "Batches"])
    for category, count in summary["category_distribution"].items():
        overview.append([category, count])

    details = workbook.create_sheet("Waste Assessments")
    columns = ["Batch ID", "Waste (kg)", "CO2 saved (kg)", "Water saved (L)", "Landfill reduced (kg)", "Recoverable (kg)", "Recyclability", "Reuse", "Recovery", "Sustainability", "Circularity", "Category", "Recommended action", "Processing method"]
    details.append(columns)
    for row in rows:
        details.append([row["batch_id"], row["quantity_kg"], row["co2_saved_kg"], row["water_saved_litres"], row["landfill_reduction_kg"], row["recoverable_material_kg"], row["recyclability_score"], row["reuse_score"], row["material_recovery_score"], row["sustainability_score"], row["circularity_score"], row["circularity_category"], row["recommended_action"], row["recommended_processing_method"]])
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor="0F766E")
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=sustainability-report.xlsx"})
